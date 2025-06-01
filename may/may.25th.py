import matplotlib.pyplot as plt
import mysql.connector
from openpyxl import load_workbook
# 加载 Excel 文件
wb = load_workbook('/Users/weiliangyu/Desktop/21统计88人.xlsx', read_only=True)
sheet = wb.active
"""
# 逐行读取数据（跳过表头）
for row in sheet.iter_rows(min_row=3, values_only=True):
    # row 是一个元组，包含当前行的所有单元格值
    data = (row[0],row[1],row[2],row[3],row[4])
"""
config={
    "host":"localhost",
    "port":3306,
    "user":"root",
    "password":"qazwsx",
    "database":"app_db",
    "auth_plugin": "mysql_native_password"
}
try:
    conn=mysql.connector.connect(**config)
    print("连接成功")
    cursor=conn.cursor()
    cursor.execute("select database();")
    result=cursor.fetchone()
    print(f"已经连接到数据库:{result[0]}")
    create_table_query="""
    CREATE TABLE IF NOT EXISTS students (
        id INT AUTO_INCREMENT PRIMARY KEY,
        name VARCHAR(50) NOT NULL COMMENT '姓名',
        student_id VARCHAR(20) NOT NULL UNIQUE COMMENT '学号',
        gender boolean COMMENT '性别',
        phone_number VARCHAR(20) COMMENT '电话号码'
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='学生基本信息表';
    """
    cursor.execute(create_table_query)
    print("学生信息表创建成功")
    insert_query = """
    INSERT INTO students (name, student_id, gender, phone_number)
    VALUES ( %s, %s,%s,%s)
    """
    # 逐行读取数据（跳过表头）
    for row in sheet.iter_rows(min_row=3, values_only=True):
        # row 是一个元组，包含当前行的所有单元格值
        data = (row[1],row[2],row[3],row[4])
        cursor.execute(insert_query, data)
    conn.commit()
    print(f"成功插入 {cursor.rowcount} 条记录")
    create_index_sql = """
    CREATE INDEX idx_name 
    ON students (name);
    """
    try:
    # 执行创建索引的 SQL
        cursor.execute(create_index_sql)
        print("索引创建成功")
    except mysql.connector.Error as err:
        print(f"索引创建失败: {err}")
    sql = f"""
    SELECT 
        gender,  -- 分组字段
        COUNT(*) AS total_people  -- 聚合函数
    FROM students
    GROUP BY gender;
    """
        # 执行查询
    cursor.execute(sql)
    # 获取所有结果
    results = cursor.fetchall() 
    print("性别统计结果：")
    genders = []
    counts = []
    for row in results:
        gender, total_people = row
        print(f"性别: {gender}, 人数: {total_people}")
        genders.append(str(gender))
        counts.append(total_people)
        plt.figure(figsize=(6, 6))
    plt.pie(counts, labels=genders, autopct='%1.2f%%')
    plt.title('different genders')
    plt.show()
except mysql.connector.Error as err:
    print(f"连接失败:{err}")
finally:
    if "conn" in locals() and conn.is_connected():
        cursor.close()
        conn.close()
        print("数据连接已关闭")